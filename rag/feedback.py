from __future__ import annotations

import threading
from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from config import FEEDBACK_XLSX_PATH
from rag.memory import _connect


_excel_lock = threading.Lock()
_HEADERS = [
    "feedback_id", "room_id", "message_id", "valoracion", "es_util",
    "pregunta_usuario", "respuesta_clara", "mensaje_creado_utc",
    "feedback_creado_utc", "feedback_actualizado_utc",
]


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _feedback_rows() -> list[dict]:
    with _connect() as con:
        rows = con.execute(
            """
            SELECT f.message_id AS feedback_id, f.room_id, f.message_id,
                   f.rating AS valoracion,
                   CASE f.rating WHEN 'useful' THEN 1 ELSE 0 END AS es_util,
                   COALESCE((SELECT u.content FROM messages u
                             WHERE u.room_id=f.room_id AND u.role='user' AND u.id < f.message_id
                             ORDER BY u.id DESC LIMIT 1), '') AS pregunta_usuario,
                   a.content AS respuesta_clara, a.created_at AS mensaje_creado_utc,
                   f.created_at AS feedback_creado_utc,
                   f.updated_at AS feedback_actualizado_utc
            FROM message_feedback f
            JOIN messages a ON a.id=f.message_id
            ORDER BY f.updated_at ASC
            """
        ).fetchall()
    return [dict(row) for row in rows]


def _export_excel() -> None:
    FEEDBACK_XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    with _excel_lock:
        if FEEDBACK_XLSX_PATH.exists():
            workbook = load_workbook(FEEDBACK_XLSX_PATH)
            sheet = workbook["Satisfacción"]
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Satisfacción"
            sheet.append(_HEADERS)
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="174C3F")
        existing = {sheet.cell(row=i, column=3).value: i for i in range(2, sheet.max_row + 1)}
        for row in _feedback_rows():
            values = [row[column] for column in _HEADERS]
            excel_row = existing.get(row["message_id"])
            if excel_row:
                for column, value in enumerate(values, start=1):
                    sheet.cell(row=excel_row, column=column, value=value)
            else:
                sheet.append(values)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        widths = {"A": 14, "B": 38, "C": 12, "D": 14, "E": 10, "F": 48, "G": 70,
                  "H": 26, "I": 26, "J": 26}
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        workbook.save(FEEDBACK_XLSX_PATH)


def save_feedback(room_id: str, message_id: int, rating: str) -> dict:
    if rating not in {"useful", "not_useful"}:
        raise ValueError("Valoración inválida.")
    now = _now()
    with _connect() as con:
        message = con.execute(
            "SELECT id FROM messages WHERE id=? AND room_id=? AND role='assistant'",
            (message_id, room_id),
        ).fetchone()
        if not message:
            raise LookupError("La respuesta no existe en esta sala.")
        con.execute(
            """INSERT INTO message_feedback(message_id,room_id,rating,created_at,updated_at)
               VALUES(?,?,?,?,?)
               ON CONFLICT(message_id) DO UPDATE SET rating=excluded.rating,updated_at=excluded.updated_at""",
            (message_id, room_id, rating, now, now),
        )
    _export_excel()
    return {"room_id": room_id, "message_id": message_id, "rating": rating}
