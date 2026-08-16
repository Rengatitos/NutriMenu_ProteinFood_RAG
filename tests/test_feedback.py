from openpyxl import load_workbook

import rag.feedback as feedback
import rag.memory as memory


def test_feedback_is_saved_and_exported(tmp_path, monkeypatch):
    db_path = tmp_path / "chats.db"
    xlsx_path = tmp_path / "feedback.xlsx"
    monkeypatch.setattr(memory, "DB_PATH", db_path)
    monkeypatch.setattr(feedback, "FEEDBACK_XLSX_PATH", xlsx_path)

    memory.init_db()
    room = memory.create_room()
    memory.save_message(room["id"], "user", "Quiero un postre")
    message_id = memory.save_message(room["id"], "assistant", "Te recomiendo una opción.")

    feedback.save_feedback(room["id"], message_id, "useful")
    feedback.save_feedback(room["id"], message_id, "not_useful")

    workbook = load_workbook(xlsx_path)
    sheet = workbook["Satisfacción"]
    assert sheet.max_row == 2
    assert sheet["B2"].value == room["id"]
    assert sheet["D2"].value == "not_useful"
    assert sheet["F2"].value == "Quiero un postre"
    assert sheet["G2"].value == "Te recomiendo una opción."
